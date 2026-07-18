"""Begleitdokumente reader: csv and xlsx supporting records become Document rows.

Files are classified by bilingual name patterns first, header sniffing second,
so a differently named or English dossier still maps. ref, entity_id, amount
and doc_date resolve through DE+EN synonym sets over the actual headers, never
by position. Every original column lands in Document.fields as strings.
The bank_confirmation and bank_statement patterns have no practice-set example
and stay dormant until such a file appears.
"""

from __future__ import annotations

import csv
import math
import os
import re
import unicodedata
import xml.etree.ElementTree as ET
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from ..contracts import Document, Dossier, SourceRef

_EXCERPT_LEN = 200
_DATE_FORMATS = ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d.%m.%y")


def _norm(s: str) -> str:
    """Casefold, transliterate umlauts, collapse separators to single spaces."""
    s = unicodedata.normalize("NFC", s or "").casefold()
    for a, b in (("\xe4", "ae"), ("\xf6", "oe"), ("\xfc", "ue"), ("\xdf", "ss")):
        s = s.replace(a, b)
    return re.sub(r"[^a-z0-9]+", " ", s).strip()


def _has(hay_n: str, *keys: str) -> bool:
    padded = f" {hay_n} "
    return any(k in padded for k in keys)


def _cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else str(v)
    if isinstance(v, datetime):
        t = (v.hour, v.minute, v.second)
        return v.date().isoformat() if t == (0, 0, 0) else v.isoformat()
    return str(v).strip()


def parse_amount(v) -> Decimal | None:
    """Handles native numbers, 1.234,56 and 1,234.56 and 1234.56 and (1,234.56)."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, Decimal):
        return v
    if isinstance(v, float) and not math.isfinite(v):
        return None
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    s = str(v).strip().replace("\xa0", "").replace(" ", "")
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    s = re.sub(r"^[a-zA-Z€$\xa3]+|[a-zA-Z€$\xa3%]+$", "", s)
    if not s or not re.search(r"\d", s):
        return None
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        head, _, tail = s.rpartition(",")
        if s.count(",") == 1 and len(tail) <= 2:
            s = head + "." + tail
        else:
            s = s.replace(",", "")
    elif s.count(".") > 1:
        s = s.replace(".", "")
    try:
        d = Decimal(s)
    except InvalidOperation:
        return None
    return -d if neg else d


def parse_date(v) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v or "").strip()
    if not s:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# ------------------------------------------------------------- classification

# Name patterns, ordered most specific first. Spaces around short keys make
# them token matches against the space-normalized name.
_NAME_RULES: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("goods_receipt", ("wareneingang", "goods receipt", "goodsreceipt", " grn ")),
    ("goods_issue", ("warenausgang", "goods issue", "goodsissue", "despatch",
                     "dispatch", "delivery note", "lieferschein")),
    ("approval", ("freigabe", "approval", "genehmigung")),
    ("master_change", ("stammdaten", "master data", "masterdata", "masterchange")),
    ("next_period_posting", ("folgeperiode", "next period", "nextperiod",
                             "subsequent period", "folgejahr")),
    ("open_item", ("op liste", "opliste", "offene posten", "offeneposten",
                   "open item", "openitem", " aging ")),
    ("trial_balance", ("saldenliste", "trial balance", "trialbalance",
                       "summen und salden", " susa ")),
    ("permission", ("berechtigung", "permission", "authorization", "authorisation",
                    "access right", "user access", "user role")),
    ("reconciliation_statement", ("abstimmung", "reconciliation", "ueberleitung")),
    ("credit_limit", ("kreditlimit", "credit limit", "creditlimit")),
    ("shareholder", ("gesellschafter", "beteiligung", "shareholder", "ownership",
                     "anteilseigner")),
    ("bank_confirmation", ("bankbestaetigung", "saldenbestaetigung", "bank confirmation",
                           "bankconfirmation", "balance confirmation",
                           "confirmation of balance")),
    ("bank_statement", ("kontoauszug", "bankauszug", "bank statement", "bankstatement",
                        "account statement", "settlement", "abrechnung", "auszahlung",
                        "payout", "disbursement")),
    ("invoice", ("fakturajournal", "faktura", "invoice", "rechnungsjournal",
                 "rechnungsausgang", "rechnungseingang", "ausgangsrechnung",
                 "eingangsrechnung", "billing")),
)

# Header signatures for files whose name matched nothing. Every group must
# match at least one column.
_SNIFF_RULES: tuple[tuple[str, tuple[tuple[str, ...], ...]], ...] = (
    ("goods_receipt", (("wareneingang", "goods receipt", " grn "),)),
    ("goods_issue", (("warenausgang", "goods issue", "despatch", "dispatch"),)),
    ("approval", (("freigeber", "approver", "approved by", "freigabestatus"),
                  ("ersteller", "creator", "created by", "entered by", "erfasst"))),
    ("master_change", (("wert alt", "wert neu", "old value", "new value",
                        "geaendert", "changed by"), ("konto", "account"))),
    ("permission", (("benutzer", "user"),
                    ("berechtigung", "permission", "zahlungslauf", "payment run",
                     "systemadmin", "freigeben", "approve"))),
    ("credit_limit", (("kreditlimit", "credit limit"),
                      ("debitor", "kunde", "customer", "kreditor", "vendor",
                       "supplier", "konto", "account"))),
    ("shareholder", (("anteil", "beteiligung", "share", "ownership", "stake"),
                     ("name",))),
    ("bank_confirmation", (("iban", " bic ", "bank"),
                           ("bestaetigung", "confirmation", "confirmed"))),
    ("bank_statement", (("iban", "kontoauszug", "statement", "valuta", "value date",
                         "buchungstag"),
                        ("saldo", "balance", "betrag", "amount", "umsatz"))),
    ("bank_statement", (("settlement", "abrechnung", "auszahlung", "payout",
                         "merchant", "acquirer", "chargeback"),
                        ("betrag", "amount", "value", "fee", "gebuehr"))),
    ("invoice", (("rechnungsnummer", "invoice"), ("betrag", "amount", "value"))),
    ("open_item", (("beleg", "voucher", "document"),
                   ("betrag", "amount", "saldo", "balance"), ("konto", "account"))),
    ("trial_balance", (("soll", "debit"), ("haben", "credit"))),
    ("subledger_balance", (("saldo", "balance"), ("konto", "account"))),
)

_VENDORISH = ("kreditor", "lieferant", "vendor", "supplier", "creditor", "payable",
              "eingangsrechnung", "purchase")
_CUSTOMERISH = ("debitor", "kunde", "customer", "debtor", "client", "receivable",
                "ausgangsrechnung", "sales")
_BALANCE_KINDS = {"open_item", "subledger_balance", "trial_balance"}


def _kind_from_name(name_n: str) -> str | None:
    for kind, pats in _NAME_RULES:
        if _has(name_n, *pats):
            return kind
    return None


def _sniff_kind(col_norms: list[str]) -> str | None:
    for kind, groups in _SNIFF_RULES:
        if all(any(_has(n, *group) for n in col_norms) for group in groups):
            return kind
    return None


def _split_invoice(name_n: str, cols: list[str]) -> str:
    hay = [name_n] + [_norm(c) for c in cols]
    if any(_has(h, *_VENDORISH) for h in hay):
        return "purchase_invoice"
    return "sales_invoice"


def _refine_balance(kind: str, cols: list[str]) -> str:
    ns = [_norm(c) for c in cols]
    toks = set()
    for n in ns:
        toks.update(n.split())
    if any(_has(n, "kontenart", "account type") for n in ns) or (
        {"soll", "debit"} & toks and {"haben", "credit"} & toks
    ):
        return "trial_balance"
    if any(_has(n, "beleg", "voucher", "document") for n in ns) and any(
        _has(n, "betrag", "amount", "value") for n in ns
    ):
        return "open_item"
    if any(_has(n, "saldo", "balance") for n in ns):
        return "subledger_balance"
    return kind


def _resolve_kind(name_n: str, cols: list[str]) -> str | None:
    kind = _kind_from_name(name_n) or _sniff_kind([_norm(c) for c in cols])
    if kind == "invoice":
        kind = _split_invoice(name_n, cols)
    if kind in _BALANCE_KINDS:
        kind = _refine_balance(kind, cols)
    return kind


# ------------------------------------------------------------- field mapping

_ENTITY_KEYS = ("kreditor", "lieferant", "vendor", "supplier", "creditor",
                "debitor", "kunde", "customer", "debtor", "client",
                "konto", "account")
_USER_KEYS = ("benutzer", "user", "ersteller", "erfasser")
_ENTITY_AVOID = ("name", "bezeichnung", "gegen", "counter", "offset", "text",
                 "datum", "date", "gruppe", "group", "typ", "type", "art")
_AMOUNT_KEYS = ("betrag", "amount", "value", "saldo", "balance", "summe",
                "total", "limit", "umsatz", "turnover")
_AMOUNT_AVOID = ("datum", "date", "waehrung", "currency", "status", "prozent",
                 "percent", "anzahl", "count", "text", "art", "typ", "type")
_DATE_KEYS = ("belegdatum", "document date", "doc date", "invoice date",
              "fakturadatum", "buchungsdatum", "posting date", "booking date",
              "valuta", "value date", "datum", "date")
_DATE_AVOID = ("zeit", "time", "uhr")
_REF_STRONG = ("nummer", "number", "beleg", "rechnung", "invoice", "voucher",
               "receipt", "reference", "faktura")
_REF_TOKENS = {"nr", "no", "id", "ref"}
_REF_WEAK = ("konto", "account", "debitor", "kreditor", "kunde", "customer",
             "vendor", "supplier", "lieferant", "benutzer", "user")
_DATEY = ("datum", "date")


def _pick(cols: list[str], keys: tuple[str, ...], avoid: tuple[str, ...]) -> str | None:
    normed = []
    for c in cols:
        n = _norm(c)
        if any(a in n for a in avoid):
            continue
        normed.append((c, n))
    for key in keys:
        for c, n in normed:
            if key in n:
                return c
    return None


def _pick_ref(cols: list[str]) -> str | None:
    """Leftmost identifier-looking column; the document's own id usually leads."""
    for c in cols:
        n = _norm(c)
        if any(d in n for d in _DATEY):
            continue
        if any(k in n for k in _REF_STRONG) or _REF_TOKENS & set(n.split()):
            return c
    for c in cols:
        n = _norm(c)
        if any(d in n for d in _DATEY):
            continue
        if any(k in n for k in _REF_WEAK):
            return c
    return None


class _Mapper:
    def __init__(self, kind: str, cols: list[str]):
        entity_keys = (_USER_KEYS + _ENTITY_KEYS
                       if kind in ("permission", "approval")
                       else _ENTITY_KEYS + _USER_KEYS)
        self.cols = cols
        self.idx = {c: i for i, c in enumerate(cols)}
        self.c_ref = _pick_ref(cols)
        self.c_entity = _pick(cols, entity_keys, _ENTITY_AVOID)
        self.c_amount = _pick(cols, _AMOUNT_KEYS, _AMOUNT_AVOID)
        self.c_date = _pick(cols, _DATE_KEYS, _DATE_AVOID)

    def get(self, col: str | None, values: list):
        if not col:
            return None
        i = self.idx[col]
        return values[i] if i < len(values) else None


def _uniq(names: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    out = []
    for i, name in enumerate(names):
        name = name.strip() or f"col{i + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name}.{seen[name]}"
        else:
            seen[name] = 1
        out.append(name)
    return out


def _make_doc(kind: str, m: _Mapper, values: list, rel: str, line_no: int,
              sheet: str | None, raw: str | None) -> Document | None:
    cells = [_cell(v) for v in values]
    filled = sum(1 for c in cells if c)
    if filled == 0 or (filled == 1 and len(m.cols) >= 3):
        return None  # blank row or in-band section label
    fields = {}
    for i, c in enumerate(m.cols):
        fields[c] = cells[i] if i < len(cells) else ""
    for j in range(len(m.cols), len(cells)):
        fields[f"extra{j - len(m.cols) + 1}"] = cells[j]
    ref = _cell(m.get(m.c_ref, values)) or next(c for c in cells if c)
    excerpt = raw if raw is not None else ";".join(cells)
    return Document(
        kind=kind,
        ref=ref,
        source=SourceRef(file=rel, line=line_no, sheet=sheet,
                         excerpt=excerpt[:_EXCERPT_LEN]),
        entity_id=_cell(m.get(m.c_entity, values)) or None,
        doc_date=parse_date(m.get(m.c_date, values)),
        amount=parse_amount(m.get(m.c_amount, values)),
        fields=fields,
    )


# ------------------------------------------------------------------ csv

def _decode(rawbytes: bytes) -> str:
    if rawbytes.startswith(b"\xef\xbb\xbf"):
        return rawbytes.decode("utf-8-sig")
    try:
        return rawbytes.decode("utf-8")
    except UnicodeDecodeError:
        return rawbytes.decode("latin-1")


_DELIMS = (";", ",", "\t", "|")


def _sniff_delim(line: str) -> str:
    counts = {d: line.count(d) for d in _DELIMS}
    best = max(_DELIMS, key=lambda d: counts[d])
    return best if counts[best] else ";"


def _iter_csv(text: str, delim: str):
    """Yield (physical_line_no, raw, fields); open-quoted rows join the next line."""
    pending, start = "", 0
    for i, line in enumerate(text.split("\n"), 1):
        raw = line.rstrip("\r")
        if pending:
            pending += "\n" + raw
        else:
            pending, start = raw, i
        if pending.count('"') % 2 == 1:
            continue
        if pending.strip():
            yield start, pending, next(csv.reader([pending], delimiter=delim, quotechar='"'))
        pending = ""
    if pending.strip():
        yield start, pending, next(csv.reader([pending.replace('"', "")], delimiter=delim))


def _load_csv(path: Path, rel: str, dossier: Dossier) -> None:
    text = _decode(path.read_bytes())
    first_line = next((ln for ln in text.splitlines() if ln.strip()), "")
    if not first_line:
        dossier.unparsed.append((rel, "empty file"))
        return
    delim = _sniff_delim(first_line)
    name_n = _norm(path.stem)
    kind, mapper = None, None
    for line_no, raw, fields in _iter_csv(text, delim):
        if mapper is None:  # first logical row is the header
            cols = _uniq([f.strip() for f in fields])
            kind = _resolve_kind(name_n, cols)
            if kind is None:
                dossier.unparsed.append((rel, f"unrecognized table, headers {cols[:8]}"))
                return
            mapper = _Mapper(kind, cols)
            continue
        doc = _make_doc(kind, mapper, fields, rel, line_no, None, raw)
        if doc:
            dossier.documents.append(doc)


# ------------------------------------------------------------------ xlsx

_KV_RECON = ("ueberleitung", "abstimmung", "reconciliation", "bridge")


def _kv_docs(kind: str, rows: list, rel: str, sheet: str, dossier: Dossier) -> None:
    """Key/value sheet: label cells left, one value per row, title row skipped."""
    first = True
    for row_no, vals in rows:
        cells = [(i, v) for i, v in enumerate(vals) if _cell(v)]
        if not cells:
            continue
        label = " ".join(_cell(v) for _, v in cells if isinstance(v, str)).strip()
        value = next((v for _, v in reversed(cells) if not isinstance(v, str)), None)
        amount = parse_amount(value)
        if amount is None and len(cells) > 1:
            value = cells[-1][1]
            amount = parse_amount(value)
        if first:
            first = False
            if amount is None and len(cells) == 1:
                continue  # title banner
        doc_date = None
        for _, v in cells:
            if isinstance(v, str):
                doc_date = parse_date(v)
                if doc_date:
                    break
        excerpt = ";".join(_cell(v) for v in vals).strip(";")
        dossier.documents.append(Document(
            kind=kind,
            ref=(label or _cell(cells[0][1]))[:120],
            source=SourceRef(file=rel, line=row_no, sheet=sheet,
                             excerpt=excerpt[:_EXCERPT_LEN]),
            doc_date=doc_date,
            amount=amount,
            fields={"label": label, "value": _cell(value)},
        ))


def _load_sheet(ws, rel: str, name_n: str, dossier: Dossier) -> None:
    rows = [(i, list(vals)) for i, vals in enumerate(ws.iter_rows(values_only=True), 1)]
    header_at = None
    for row_no, vals in rows[:6]:
        if sum(1 for v in vals if isinstance(v, str) and v.strip()) >= 3:
            header_at = row_no
            break
    if header_at is None:
        kind = ("reconciliation_statement" if _has(_norm(ws.title), *_KV_RECON)
                else _kind_from_name(name_n))
        if kind == "invoice":
            kind = _split_invoice(name_n, [])
        if kind is None:
            if rows:
                dossier.unparsed.append((rel, f"sheet '{ws.title}': no header row found"))
            return
        _kv_docs(kind, rows, rel, ws.title, dossier)
        return
    cols = _uniq([_cell(v) for v in rows[header_at - 1][1]])
    kind = _resolve_kind(name_n, cols)
    if kind is None:
        dossier.unparsed.append((rel, f"sheet '{ws.title}': unrecognized table, headers {cols[:8]}"))
        return
    mapper = _Mapper(kind, cols)
    for row_no, vals in rows[header_at:]:
        doc = _make_doc(kind, mapper, vals, rel, row_no, ws.title, None)
        if doc:
            dossier.documents.append(doc)


def _load_xlsx(path: Path, rel: str, dossier: Dossier) -> None:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        name_n = _norm(path.stem)
        for ws in wb.worksheets:
            try:
                _load_sheet(ws, rel, name_n, dossier)
            except Exception as e:
                dossier.unparsed.append((rel, f"sheet '{ws.title}': {e}"))
    finally:
        wb.close()


# ------------------------------------------------------------------ entry

def _gdpdu_urls(root: Path) -> set[Path]:
    """Files claimed by a GDPdU index.xml; gdpdu.py owns those."""
    urls: set[Path] = set()
    for dirpath, _dirnames, filenames in os.walk(root, followlinks=True):
        if "index.xml" not in filenames:
            continue
        try:
            tree = ET.parse(Path(dirpath) / "index.xml")
            for t in tree.getroot().iter("Table"):
                u = t.findtext("URL")
                if u:
                    urls.add((Path(dirpath) / u).resolve())
        except Exception:
            continue
    return urls


def _walk(root: Path):
    for dirpath, dirnames, filenames in os.walk(root, followlinks=True):
        dirnames.sort()
        for fn in sorted(filenames):
            if fn.startswith(("~$", ".")):
                continue
            if fn.lower().endswith((".csv", ".xlsx")):
                yield Path(dirpath) / fn


def load_begleit(root: Path, dossier: Dossier) -> None:
    """Read every csv/xlsx under root into Document rows, one per data row."""
    root = Path(root)
    claimed = _gdpdu_urls(root)
    for path in _walk(root):
        rel = str(path.relative_to(root))
        if path.resolve() in claimed:
            continue
        try:
            if path.suffix.lower() == ".csv":
                _load_csv(path, rel, dossier)
            else:
                _load_xlsx(path, rel, dossier)
        except Exception as e:
            dossier.unparsed.append((rel, f"read failed: {e}"))
